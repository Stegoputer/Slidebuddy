from pathlib import Path

def parse_excel(file_path: Path) -> str:
    """Parse Excel/CSV files into readable text format."""
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return _parse_csv(file_path)
    elif suffix in (".xlsx", ".xls", ".xlsm"):
        return _parse_xlsx(file_path)
    elif suffix == ".tsv":
        return _parse_csv(file_path, delimiter="\t")
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def _parse_csv(file_path: Path, delimiter: str = ",") -> str:
    import csv

    rows = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            rows.append(" | ".join(row))
    return "\n".join(rows)


def _parse_xlsx(file_path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    sheets_text = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cell_values = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in cell_values):
                rows.append(" | ".join(cell_values))
        if rows:
            sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))

    wb.close()
    return "\n\n".join(sheets_text)
